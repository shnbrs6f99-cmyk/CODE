from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from .domain import CalculationResult, InterestRules
from .validation import ValidationMessage


class ExcelReportGenerator:
    HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
    WHITE_BOLD = Font(color="FFFFFF", bold=True)
    BOLD = Font(bold=True)
    THIN = Border(*(Side(style="thin", color="B7B7B7") for _ in range(4)))
    MONEY = '#,##0.00;[Red]-#,##0.00'
    DATE = "dd-mmm-yyyy"

    def generate(
        self,
        result: CalculationResult,
        rules: InterestRules,
        validations: list[ValidationMessage],
        output: Path,
    ) -> Path:
        output.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        summary = wb.active
        summary.title = "Summary"
        self._summary(summary, result, rules)
        self._transactions(wb.create_sheet("Transactions"), result)
        self._allocations(wb.create_sheet("FIFO Allocations"), result)
        self._interest(wb.create_sheet("Interest Calculation"), result)
        self._validation(wb.create_sheet("Validation"), validations, result)
        wb.properties.title = f"Interest Statement - {result.statement.customer_name}"
        wb.properties.creator = "Interest Statement Generator Pro"
        wb.properties.created = datetime.now()
        wb.save(output)
        return output

    def _summary(self, ws, result: CalculationResult, rules: InterestRules) -> None:
        ws.merge_cells("A1:F1")
        ws["A1"] = "INTEREST STATEMENT"
        ws["A1"].fill = self.HEADER_FILL
        ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center")
        rows = [
            ("Customer", result.statement.customer_name),
            ("Source file", result.statement.source_file.name),
            ("Statement period", f"{result.statement.period_start or '-'} to {result.statement.period_end or '-'}"),
            ("Annual interest rate", float(rules.annual_rate) / 100),
            ("Credit period", rules.credit_period_days),
            ("Opening balance", float(result.statement.opening_balance)),
            ("Closing balance", float(result.statement.closing_balance)),
            ("Outstanding principal", float(result.outstanding_principal)),
            ("Unallocated credits", float(result.unallocated_credits)),
            ("Total interest", float(result.total_interest)),
        ]
        for row, (label, value) in enumerate(rows, start=3):
            ws.cell(row, 1, label).font = self.BOLD
            ws.cell(row, 2, value)
            ws.cell(row, 1).fill = self.SUBHEADER_FILL
            ws.cell(row, 1).border = ws.cell(row, 2).border = self.THIN
        ws["B6"].number_format = "0.00%"
        for r in range(8, 13):
            ws.cell(r, 2).number_format = self.MONEY
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 38
        self._print_setup(ws, "A1:F14")

    def _transactions(self, ws, result: CalculationResult) -> None:
        headers = ["Date", "Voucher Type", "Voucher No.", "Narration", "Debit", "Credit", "Page"]
        self._table_header(ws, headers)
        for idx, tx in enumerate(result.statement.transactions, start=2):
            values = [tx.transaction_date, tx.voucher_type.value, tx.voucher_number, tx.narration, float(tx.debit), float(tx.credit), tx.source_page]
            for col, value in enumerate(values, start=1):
                ws.cell(idx, col, value).border = self.THIN
            ws.cell(idx, 1).number_format = self.DATE
            ws.cell(idx, 5).number_format = ws.cell(idx, 6).number_format = self.MONEY
        self._finish_table(ws, [14, 18, 18, 48, 16, 16, 10])

    def _allocations(self, ws, result: CalculationResult) -> None:
        headers = ["Charge Voucher", "Charge Date", "Payment Voucher", "Payment Date", "Allocated Amount"]
        self._table_header(ws, headers)
        for idx, item in enumerate(result.allocations, start=2):
            values = [item.charge_voucher, item.charge_date, item.payment_voucher, item.payment_date, float(item.amount)]
            for col, value in enumerate(values, start=1):
                ws.cell(idx, col, value).border = self.THIN
            ws.cell(idx, 2).number_format = ws.cell(idx, 4).number_format = self.DATE
            ws.cell(idx, 5).number_format = self.MONEY
        self._finish_table(ws, [20, 14, 20, 14, 20])

    def _interest(self, ws, result: CalculationResult) -> None:
        headers = ["Voucher", "Type", "Transaction Date", "Due Date", "Settlement/Cut-off", "Principal", "Days", "Rate", "Interest"]
        self._table_header(ws, headers)
        for idx, line in enumerate(result.interest_lines, start=2):
            values = [line.voucher_number, line.voucher_type.value, line.transaction_date, line.due_date, line.settled_date, float(line.principal), line.days_overdue, float(line.annual_rate) / 100, float(line.interest)]
            for col, value in enumerate(values, start=1):
                ws.cell(idx, col, value).border = self.THIN
            for col in (3, 4, 5):
                ws.cell(idx, col).number_format = self.DATE
            ws.cell(idx, 6).number_format = ws.cell(idx, 9).number_format = self.MONEY
            ws.cell(idx, 8).number_format = "0.00%"
        total_row = len(result.interest_lines) + 2
        ws.cell(total_row, 8, "Total").font = self.BOLD
        ws.cell(total_row, 9, float(result.total_interest)).number_format = self.MONEY
        ws.cell(total_row, 9).font = self.BOLD
        self._finish_table(ws, [18, 18, 16, 16, 18, 16, 10, 12, 16])

    def _validation(self, ws, validations, result: CalculationResult) -> None:
        self._table_header(ws, ["Severity", "Code", "Message"])
        rows = validations or []
        for idx, item in enumerate(rows, start=2):
            for col, value in enumerate([item.severity.value.upper(), item.code, item.message], start=1):
                ws.cell(idx, col, value).border = self.THIN
        start = len(rows) + 4
        ws.cell(start, 1, "Parser").font = self.BOLD
        ws.cell(start, 2, result.statement.parser_name)
        ws.cell(start + 1, 1, "Confidence").font = self.BOLD
        ws.cell(start + 1, 2, result.statement.confidence).number_format = "0%"
        self._finish_table(ws, [14, 24, 90])

    def _table_header(self, ws, headers: list[str]) -> None:
        for col, text in enumerate(headers, start=1):
            cell = ws.cell(1, col, text)
            cell.fill = self.HEADER_FILL
            cell.font = self.WHITE_BOLD
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.THIN
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    def _finish_table(self, ws, widths: list[int]) -> None:
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.sheet_view.showGridLines = False
        self._print_setup(ws, ws.dimensions)

    @staticmethod
    def _print_setup(ws, area: str) -> None:
        ws.print_area = area
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5, header=0.2, footer=0.2)
        ws.oddFooter.center.text = "Page &P of &N"
