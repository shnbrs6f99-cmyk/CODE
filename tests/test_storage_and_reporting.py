from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from interest_statement_pro.domain import CalculationResult, InterestRules, LedgerStatement
from interest_statement_pro.reporting import ExcelReportGenerator
from interest_statement_pro.storage import Database


def test_database_settings_and_integrity(tmp_path: Path):
    db = Database(tmp_path / "app.db")
    db.save_setting("theme", "dark")
    db.upsert_customer("Acme", rate="18", credit_days=30)
    assert db.load_setting("theme") == "dark"
    assert db.customers()[0]["name"] == "Acme"
    assert db.integrity_check()


def test_excel_workbook_contains_expected_sheets(tmp_path: Path):
    statement = LedgerStatement(
        "Acme", date(2026, 1, 1), date(2026, 1, 31), Decimal("0"), Decimal("0"), [],
        Path("ledger.pdf"), "test",
    )
    result = CalculationResult(statement, [], [], Decimal("0"), Decimal("0"), Decimal("0"))
    output = tmp_path / "statement.xlsx"
    ExcelReportGenerator().generate(result, InterestRules(), [], output)
    workbook = load_workbook(output, read_only=True)
    assert workbook.sheetnames == [
        "Summary", "Transactions", "FIFO Allocations", "Interest Calculation", "Validation"
    ]
